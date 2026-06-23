import streamlit as st
from openpyxl import load_workbook
import tempfile
import pandas as pd
import pdfplumber
import unicodedata
from io import BytesIO

import matching_logic as ml

st.set_page_config(
    page_title="変状図・点検表 整合性チェック",
    layout="wide",
)

st.title("変状図・点検表 整合性チェックシステム")

col_up1, col_up2 = st.columns(2)
with col_up1:
    excel_file = st.file_uploader("点検表Excelをアップロード", type=["xlsx"])
with col_up2:
    pdf_file = st.file_uploader("変状図PDFをアップロード", type=["pdf"])

with st.expander("⚙️ 判定ロジックの設定ファイルを変更する（上級者向け）"):
    st.caption(
        "通常は items_config.json を直接編集してください。"
        "別の設定ファイルを一時的に試したい場合は、ここにアップロードしてください。"
    )
    custom_config_file = st.file_uploader(
        "カスタム設定ファイル（config.json）", type=["json"], key="custom_config"
    )
    if custom_config_file is not None:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".json") as tmp_cfg:
            tmp_cfg.write(custom_config_file.read())
            custom_config_path = tmp_cfg.name
        ml.reload_config(custom_config_path)
        st.success("カスタム設定を読み込みました。")
    else:
        ml.reload_config()  # items_config.json（デフォルト）に戻す

# 判定ロジックの定数はすべて matching_logic（設定ファイル）から取得する。
# 点検項目・行番号・PDF検索キーワード等を変更したい場合は、コードではなく
# items_config.json を編集すること。
ITEM_ROWS = ml.ITEM_ROWS
ADDRESS_ITEM_NAME = ml.ADDRESS_ITEM_NAME
NO_PDF_ITEMS = ml.NO_PDF_ITEMS
EDGE_STONE_ITEM_NAME = ml.EDGE_STONE_ITEM_NAME
LEVEL_SET = ml.LEVEL_SET


# ============================================================
# 1. データ処理（画面表示は後でまとめて行う）
# ============================================================

excel_df = None
pdf_df = None
sheet_names = []
notices = []  # (level, message) を処理中にここへ集めて、後でまとめて表示する

# PDFのページ番号で直接ページ送りできるようにするための情報
total_pdf_pages = 0
pdf_page_to_no = {}      # {PDFページ番号: 検出された換気口No}
pdf_page_texts = {}      # {PDFページ番号: そのページの抽出テキスト（確認・デバッグ用）}


def add_notice(level, message):
    notices.append((level, message))


if excel_file is not None:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(excel_file.read())
        excel_path = tmp.name

    wb = load_workbook(excel_path, data_only=True)
    sheet_names = sorted(wb.sheetnames, key=lambda x: int(x))

    excel_rows = []
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        for item, row_num in ITEM_ROWS.items():
            excel_rows.append({
                "換気口No": sheet_name,
                "点検項目": item,
                "Excel_変状有無": ws.cell(row=row_num, column=2).value,
                "Excel_形状": ws.cell(row=row_num, column=3).value,
            })

    excel_df = pd.DataFrame(excel_rows)

if pdf_file is not None and excel_file is not None:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp.write(pdf_file.read())
        pdf_path = tmp.name

    pdf_rows = []

    with pdfplumber.open(pdf_path) as pdf:
        total_pdf_pages = len(pdf.pages)
        all_pages = {}  # {ページ番号: Pageオブジェクト}（全ページ分）
        page_kanki_nos = {}  # {ページ番号: 検出された換気口No（Noneなら未検出）}

        for page_index, page in enumerate(pdf.pages, start=1):
            all_pages[page_index] = page

            page_text = page.extract_text() or ""
            pdf_page_texts[page_index] = page_text

            page_kanki_nos[page_index] = ml.extract_kanki_no(page)

        # 同じ換気口Noが複数ページに分かれている場合（例：換気口内部の図と
        # 換気口路面部の図が別ページになっている等）もあるため、上書きせずに
        # ページ番号を全て集める（matching_logic.build_no_to_pages に委譲）。
        pdf_no_to_pages, pdf_page_to_no_result, no_detected_pages = ml.build_no_to_pages(
            page_kanki_nos
        )
        pdf_page_to_no.update(pdf_page_to_no_result)

        if no_detected_pages:
            add_notice("warning", f"換気口Noを検出できなかったPDFページ：{no_detected_pages}")

        multi_page_nos = {no: pages for no, pages in pdf_no_to_pages.items() if len(pages) > 1}
        if multi_page_nos:
            summary = "、".join(
                f"換気口No「{no}」は{', '.join(str(p) + 'ページ目' for p in pages)}"
                for no, pages in multi_page_nos.items()
            )
            add_notice(
                "info",
                f"以下の換気口Noは複数ページに分かれて記載されているため、まとめて判定します：{summary}"
            )

        page_words_cache = {}

        def get_page_words(target_page_index):
            if target_page_index not in page_words_cache:
                target_page = all_pages.get(target_page_index)
                try:
                    # 大きな見出し文字は文字同士の間隔が広いことがあり、デフォルトの
                    # 許容距離では「ひび割れ」のような項目名が文字ごとに分割されて
                    # しまうことがある。x_toleranceを広げて結合しやすくする。
                    page_words_cache[target_page_index] = (
                        target_page.extract_words(x_tolerance=15, y_tolerance=5)
                        if target_page else []
                    )
                except Exception:
                    page_words_cache[target_page_index] = []
            return page_words_cache[target_page_index]

        missing_in_pdf = []
        address_not_detected = []

        for sheet_name in sheet_names:
            page_indices = pdf_no_to_pages.get(str(sheet_name))

            if not page_indices:
                missing_in_pdf.append(sheet_name)
                continue

            page_index = page_indices[0]  # 一覧表示用の代表ページ番号

            text = " ".join(pdf_page_texts.get(pi, "") for pi in page_indices)
            text = text.replace("　", " ")

            # 項目名とレベル文字（A/B/C/S等）の位置情報を取得する。
            # このPDFは表ではなく図面上に項目名とレベル文字が個別に配置されているため、
            # ページ全体で最初に見つかったレベルではなく、項目名と位置が近いものだけを使う。
            words = []
            for pi in page_indices:
                page_word_list = get_page_words(pi)
                for w in page_word_list:
                    if "_page_index" not in w:
                        w["_page_index"] = pi
                words.extend(page_word_list)

            level_words = [
                w for w in words
                if unicodedata.normalize("NFKC", w.get("text", "")).strip() in LEVEL_SET
            ]

            for item in ITEM_ROWS.keys():
                # 判定ロジック本体はすべて matching_logic.evaluate_item() に
                # 委譲している（UIとロジックを分離し、テストしやすくするため）。
                result = ml.evaluate_item(item, text, words, level_words)

                if item == ADDRESS_ITEM_NAME and not result["address_detected"]:
                    address_not_detected.append(sheet_name)

                pdf_rows.append({
                    "PDFページ番号": ",".join(str(p) for p in page_indices),
                    "換気口No": sheet_name,
                    "点検項目": item,
                    "PDF_変状有無": result["pdf_status"],
                    "PDF_形状": result["pdf_shape"],
                })

        if missing_in_pdf:
            add_notice("warning", f"PDFに対応ページが見つからなかった換気口No（Excel基準）：{missing_in_pdf}")

        if address_not_detected:
            add_notice("warning", f"PDFから地番・目標物のテキストを検出できなかった換気口No：{address_not_detected}")

        extra_in_pdf = sorted(
            set(pdf_no_to_pages.keys()) - {str(s) for s in sheet_names}
        )
        if extra_in_pdf:
            add_notice("info", f"Excelには無いがPDFに存在する換気口No：{extra_in_pdf}")

    pdf_df = pd.DataFrame(pdf_rows)

elif pdf_file is not None and excel_file is None:
    add_notice("warning", "PDFを比較するには、先にExcel点検表もアップロードしてください。")


# ============================================================
# 2. 画面表示（整合性の判定を最優先・最大で表示する）
# ============================================================

if excel_df is not None and pdf_df is not None:

    result_df = pd.merge(
        excel_df,
        pdf_df,
        on=["換気口No", "点検項目"],
        how="inner"
    ).fillna("")

    is_excluded_mask = result_df.apply(
        lambda row: ml.is_excluded(row["点検項目"], row["Excel_変状有無"], row["PDF_変状有無"]),
        axis=1,
    )

    excluded_df = result_df[is_excluded_mask]
    compare_df = result_df[~is_excluded_mask].copy()

    compare_df["判定"] = compare_df.apply(
        lambda row: ml.judge(row["点検項目"], row["Excel_変状有無"], row["PDF_変状有無"]),
        axis=1,
    )
    ng_df = compare_df[compare_df["判定"] == "NG"]

    ok_count = len(compare_df[compare_df["判定"] == "OK"])
    ng_count = len(ng_df)
    target_count = len(compare_df)
    excluded_count = len(excluded_df)

    # ---- 最優先表示：総合判定バナー ----
    st.markdown("###")

    if ng_count == 0:
        st.markdown(
            """
            <div style="background-color:#e6f4ea;border:3px solid #34a853;
                        border-radius:14px;padding:32px;text-align:center;">
                <div style="font-size:48px;line-height:1;">✅</div>
                <div style="font-size:30px;font-weight:700;color:#188038;margin-top:8px;">
                    整合性OK
                </div>
                <div style="font-size:16px;color:#3c4043;margin-top:6px;">
                    Excel点検表とPDF変状図の内容は一致しています
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            f"""
            <div style="background-color:#fce8e6;border:3px solid #d93025;
                        border-radius:14px;padding:32px;text-align:center;">
                <div style="font-size:48px;line-height:1;">⚠️</div>
                <div style="font-size:30px;font-weight:700;color:#c5221f;margin-top:8px;">
                    不一致あり：{ng_count}件
                </div>
                <div style="font-size:16px;color:#3c4043;margin-top:6px;">
                    内容を確認してください（詳細は「不一致一覧」タブ）
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown("###")

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("比較対象", target_count)
    col2.metric("一致", ok_count)
    col3.metric("不一致", ng_count, delta=None)
    col4.metric("対象外", excluded_count)

    # ---- 注意事項（換気口Noの紐付けに関する警告など） ----
    if notices:
        with st.expander(f"⚠️ 確認事項（{len(notices)}件）", expanded=(ng_count > 0 or any(n[0] == "warning" for n in notices))):
            for level, message in notices:
                if level == "warning":
                    st.warning(message)
                else:
                    st.info(message)

    st.markdown("---")

    # ---- ダウンロード ----
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        compare_df.to_excel(writer, index=False, sheet_name="比較対象一覧")
        ng_df.to_excel(writer, index=False, sheet_name="不一致一覧")
        excluded_df.to_excel(writer, index=False, sheet_name="対象外一覧")

    st.download_button(
        label="📥 チェック結果をExcelでダウンロード",
        data=output.getvalue(),
        file_name="整合性チェック結果.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch",
    )

    st.markdown("###")

    # ---- 詳細データ（タブでまとめて表示） ----
    tab_ng, tab_all, tab_excluded, tab_raw = st.tabs(
        ["🔴 不一致一覧", "📋 比較対象一覧", "➖ 対象外一覧", "🗂 読み取り結果（Excel/PDF）"]
    )

    with tab_ng:
        if ng_count == 0:
            st.success("不一致はありません。")
        else:
            st.dataframe(ng_df, width="stretch", hide_index=True)

    with tab_all:
        st.dataframe(compare_df, width="stretch", hide_index=True)

    with tab_excluded:
        st.caption(
            "Excel側の「変状有無」が空欄、「歩道幅員」「その他」のようにそもそもPDF"
            "（変状図）に記載が無く比較できない項目、または「縁石コンクリート破損」で"
            "「縁石」の記載が無く「ひび割れ」表記のみで自動判定できない項目（要確認）です。"
        )
        st.dataframe(excluded_df, width="stretch", hide_index=True)

    with tab_raw:
        if not sheet_names:
            st.info("データがありません。")
        else:
            view_mode = st.radio(
                "ページ送りの方法",
                ["PDFページ番号で選ぶ", "換気口Noで選ぶ"],
                horizontal=True,
                index=0,
            )

            st.markdown("---")

            if view_mode == "PDFページ番号で選ぶ":
                # ---- PDFの実際のページ番号で直接めくるモード ----
                try:
                    if total_pdf_pages == 0:
                        st.info("PDFのページ情報がありません。")
                    else:
                        if "raw_pdf_page_index" not in st.session_state:
                            st.session_state.raw_pdf_page_index = 0

                        st.session_state.raw_pdf_page_index = max(
                            0, min(st.session_state.raw_pdf_page_index, total_pdf_pages - 1)
                        )

                        col_prev, col_select, col_next = st.columns([1, 4, 1])

                        with col_prev:
                            if st.button(
                                "◀ 前のページ",
                                width="stretch",
                                disabled=st.session_state.raw_pdf_page_index <= 0,
                                key="pdf_page_prev",
                            ):
                                st.session_state.raw_pdf_page_index -= 1
                                st.rerun()

                        with col_next:
                            if st.button(
                                "次のページ ▶",
                                width="stretch",
                                disabled=st.session_state.raw_pdf_page_index >= total_pdf_pages - 1,
                                key="pdf_page_next",
                            ):
                                st.session_state.raw_pdf_page_index += 1
                                st.rerun()

                        with col_select:
                            selected_page = st.selectbox(
                                "PDFのページ番号を選択",
                                options=list(range(1, total_pdf_pages + 1)),
                                index=st.session_state.raw_pdf_page_index,
                                key="pdf_page_selectbox",
                            )

                        # ボタン操作と選択ボックスの状態を同期する
                        if selected_page - 1 != st.session_state.raw_pdf_page_index:
                            st.session_state.raw_pdf_page_index = selected_page - 1
                            st.rerun()

                        st.write("")  # わずかな余白
                        st.caption(f"PDFの {selected_page} / {total_pdf_pages} ページ目を表示しています")

                        matched_no = pdf_page_to_no.get(selected_page)

                        if matched_no is None:
                            st.warning(
                                "このページからは換気口Noを検出できませんでした。"
                                "表紙・見出しページなど、点検データを含まないページの可能性があります。"
                            )
                        else:
                            st.success(f"このページの換気口No：「{matched_no}」")

                            excel_page_df = excel_df[excel_df["換気口No"] == matched_no]
                            pdf_page_df = pdf_df[pdf_df["換気口No"] == matched_no]

                            st.markdown("**Excel点検表の読み取り結果**")
                            st.dataframe(excel_page_df, width="stretch", hide_index=True)

                            st.markdown("**PDF変状図の読み取り結果**")
                            st.dataframe(pdf_page_df, width="stretch", hide_index=True)

                        with st.expander("このページから抽出した元テキストを確認（確認・デバッグ用）"):
                            page_text = pdf_page_texts.get(selected_page, "")
                            if page_text:
                                st.text(page_text)
                            else:
                                st.caption("テキストを抽出できませんでした。")
                except Exception as e:
                    st.error("このタブの表示中にエラーが発生しました。下記の内容を共有してください。")
                    st.exception(e)

            else:
                # ---- 換気口Noの昇順で選ぶモード（従来の方式） ----
                try:
                    page_order = sorted(sheet_names, key=lambda x: int(x))

                    if "raw_page_index" not in st.session_state:
                        st.session_state.raw_page_index = 0

                    # 換気口Noの数が変わった場合などにインデックスが範囲外にならないよう補正
                    st.session_state.raw_page_index = max(
                        0, min(st.session_state.raw_page_index, len(page_order) - 1)
                    )

                    col_prev, col_select, col_next = st.columns([1, 4, 1])

                    with col_prev:
                        if st.button(
                            "◀ 前へ",
                            width="stretch",
                            disabled=st.session_state.raw_page_index <= 0,
                            key="no_page_prev",
                        ):
                            st.session_state.raw_page_index -= 1
                            st.rerun()

                    with col_next:
                        if st.button(
                            "次へ ▶",
                            width="stretch",
                            disabled=st.session_state.raw_page_index >= len(page_order) - 1,
                            key="no_page_next",
                        ):
                            st.session_state.raw_page_index += 1
                            st.rerun()

                    with col_select:
                        selected_no = st.selectbox(
                            "換気口Noを選択",
                            options=page_order,
                            index=st.session_state.raw_page_index,
                            key="raw_page_selectbox",
                        )

                    if page_order.index(selected_no) != st.session_state.raw_page_index:
                        st.session_state.raw_page_index = page_order.index(selected_no)
                        st.rerun()

                    pdf_pages_for_no = pdf_no_to_pages.get(str(selected_no), [])
                    if pdf_pages_for_no:
                        pages_text = "、".join(f"{p}ページ目" for p in pdf_pages_for_no)
                        page_info = f"（PDFの{pages_text}）"
                    else:
                        page_info = "（PDFに対応ページなし）"

                    st.caption(
                        f"{st.session_state.raw_page_index + 1} / {len(page_order)} 件目　"
                        f"換気口No「{selected_no}」　{page_info}"
                    )

                    excel_page_df = excel_df[excel_df["換気口No"] == selected_no]
                    pdf_page_df = pdf_df[pdf_df["換気口No"] == selected_no]

                    st.markdown("**Excel点検表の読み取り結果**")
                    st.dataframe(excel_page_df, width="stretch", hide_index=True)

                    st.markdown("**PDF変状図の読み取り結果**")
                    st.dataframe(pdf_page_df, width="stretch", hide_index=True)
                except Exception as e:
                    st.error("このタブの表示中にエラーが発生しました。下記の内容を共有してください。")
                    st.exception(e)

elif excel_df is not None and pdf_file is None:
    st.info("PDFもアップロードすると、整合性チェックが行われます。")
    with st.expander("Excel点検表の読み取り結果を確認"):
        st.dataframe(excel_df, width="stretch")

elif notices:
    for level, message in notices:
        if level == "warning":
            st.warning(message)
        else:
            st.info(message)
